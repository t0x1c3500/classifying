import os
import csv
from openpyxl import Workbook


directory_path = './list_16/results'
list_results = []

for filename in os.listdir(directory_path):
    if filename.endswith('.txt'):
        file_path = os.path.join(directory_path, filename)

        with open(file_path, 'r') as file:
            content = file.read().strip()

            if content.isdigit():
                list_results.append({
                    'StoreName': f"{filename.replace('.txt', '')}",
                    'StoreUrl': f"https://{filename.replace('.txt', '')}/collections/all?sort_by=created-descending",
                    'AdsUrl': f"https://www.pipiads.com/ad-search?"
                              f"search_type=1&extend_keywords=%5B%7B%22type%22%3A1,"
                              f"%22keyword%22%3A%22{filename.replace('.txt', '')}%22%7D%5D&sort=2"
                              f"&sort_type=desc&current_page=1",
                    'FbAdsUrl': f"https://www.facebook.com/ads/library/?active_status=all&ad_type=all&country=ALL&"
                                f"media_type=all&q=%22{filename.replace('.txt', '')}%22&search_type=keyword_exact_phrase",
                    'DropPercentage': int(content),
                    'isDropShip': ''
                })
            else:
                print(f"Value Error : {filename.replace('.txt', '')}")


workbook = Workbook()
worksheet = workbook.active

if isinstance(list_results, list) and len(list_results) > 0:
    headers = list(list_results[0].keys())
    worksheet.append(headers)

    for row_data in list_results:
        row_index = worksheet.max_row + 1
        for col_index, key in enumerate(headers, start=1):
            value = row_data.get(key, "")
            cell = worksheet.cell(row=row_index, column=col_index)

            if key in ["StoreUrl", "AdsUrl", "FbAdsUrl"] and isinstance(value, str) and value.startswith("http"):
                cell.value = value
                cell.hyperlink = value
                cell.style = "Hyperlink"
            else:
                cell.value = value

workbook.save("list_16-results.xlsx")